#!/usr/bin/env python3
"""Danh so lai: MOI CAN NHA mot so, hau to phan biet tang trong can do.

    bietthu_002.jpg      mat bang tret cua can so 2
    bietthu_002_t1.jpg   tang 1 cua CUNG can do
    bietthu_002_t2.jpg   tang 2 cua CUNG can do

Truoc day moi ANH mot so, nen ba ban ve cua cung mot can mang ba so khac nhau
va khong nhin ra chung thuoc cung mot cong trinh.

  python danh_so_theo_can.py            # xem truoc
  python danh_so_theo_can.py --that     # thuc hien
"""
from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict

from openpyxl import load_workbook

import assess
import config as C
import main as M

# Thu tu tang de sap trong mot can: tret truoc, roi tang so, roi tang dac biet
THU_TU = {"": 0, "_th": -1, "_tl": 5, "_tst": 90, "_tap": 91, "_tum": 92}


def khoa_sap(hau_to: str) -> float:
    if hau_to in THU_TU:
        return THU_TU[hau_to]
    if hau_to.startswith("_t") and hau_to[2:].isdigit():
        return int(hau_to[2:])
    return 99


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--that", action="store_true", help="thuc hien that")
    args = ap.parse_args()

    if not os.path.exists(M.EXCEL):
        sys.exit(f"Khong thay {M.EXCEL}")

    wb = load_workbook(M.EXCEL)
    ws = wb["dataset"]
    base = len(C.EXCEL_COLUMNS)
    prefix = C.HOUSE_TYPE + "_"

    # 1. Gom dong theo CAN NHA.
    # Khoa = (trang du an, ten can rut tu ten file). Chi dung trang du an la
    # khong du: nguon kieu trang tong hop co mot trang chua hang chuc can.
    theo_can: dict[tuple, list] = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        ten = str(row[0].value or "")
        if not ten.startswith(prefix):
            continue                       # mau 'loai' dat ten theo pHash
        ten_goc = str(row[base + 1].value or "").rsplit("/", 1)[-1]
        khoa = (str(row[2].value or "khong_ro"), assess.khoa_can_nha(ten_goc))
        theo_can[khoa].append((row, ten, ten_goc))

    # 2. Giu thu tu can theo so nho nhat hien co, de xao tron it nhat
    def so_nho_nhat(ds) -> int:
        ra = []
        for _, ten, _f in ds:
            s = os.path.splitext(ten)[0][len(prefix):][:3]
            ra.append(int(s) if s.isdigit() else 9999)
        return min(ra)

    thu_tu_can = sorted(theo_can, key=lambda c: so_nho_nhat(theo_can[c]))

    doi: dict[str, str] = {}
    trung = 0
    for i, can in enumerate(thu_tu_can, 1):
        ds = theo_can[can]
        # Can nay co ban ve TRET khong? Quyet dinh cach dich so tang:
        # nguon ghi "tret, tang 2, tang 3" thi tang 2 la tang thu nhat tren tret.
        co_tret = any(assess.nhan_tang(f) == "" and C.ALT_TRET.search(
            assess._cat_chu_de(f)) for _r, _t, f in ds)

        co_hau_to = [(row, ten, assess.nhan_tang(f, co_tret)) for row, ten, f in ds]
        co_hau_to.sort(key=lambda x: khoa_sap(x[2]))

        da_dung: dict[str, int] = {}
        for row, ten, hau_to in co_hau_to:
            n = da_dung.get(hau_to, 0)
            da_dung[hau_to] = n + 1
            # Hai ban ve cung tang trong cung can -> them chu cai phan biet
            them = "" if n == 0 else chr(ord("b") + n - 1)
            if n:
                trung += 1
            ext = os.path.splitext(ten)[1]
            moi = f"{prefix}{i:03d}{hau_to}{them}{ext}"
            if moi != ten:
                doi[ten] = moi

    print(f"{len(thu_tu_can)} can nha | {sum(len(v) for v in theo_can.values())} anh")
    print(f"{len(doi)} file doi ten | {trung} anh trung tang trong cung can\n")
    for cu, moi in list(doi.items())[:12]:
        print(f"  {cu:<24} -> {moi}")
    if len(doi) > 12:
        print(f"  ... con {len(doi) - 12} file")

    if not args.that:
        print("\nChay lai voi --that de thuc hien.")
        return

    # 3. Doi ten qua buoc trung gian de khong de len file dang ton tai
    tam = []
    for i, (cu, moi) in enumerate(doi.items()):
        for d in (M.IMG_DIR, M.TAM_DIR, M.LOAI_DIR):
            p = os.path.join(d, cu)
            if os.path.exists(p):
                t = os.path.join(d, f"__dc_{i:04d}{os.path.splitext(cu)[1]}")
                os.replace(p, t)
                tam.append((t, os.path.join(d, moi)))
                break
    for t, dich in tam:
        os.replace(t, dich)

    sua = 0
    for row in ws.iter_rows(min_row=2):
        v = str(row[0].value or "")
        if v in doi:
            row[0].value = doi[v]
            sua += 1
    wb.save(M.EXCEL)
    print(f"\nDa doi {len(tam)} file | Excel sua {sua} dong")


if __name__ == "__main__":
    sys.exit(main())
