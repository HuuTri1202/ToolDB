"""Ghi Excel dung 10 cot theo muc II.5 + bao cao tien do."""
from __future__ import annotations

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config as C

FILL = {
    "dat": PatternFill("solid", fgColor="E1F5EE"),
    "tam": PatternFill("solid", fgColor="FAEEDA"),
    "loai": PatternFill("solid", fgColor="FCEBEB"),
}


def _new_book(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "dataset"
    ws.append(C.EXCEL_COLUMNS + ["_trang_thai", "_url_nguon", "_phash", "_ma_du_an"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col, width in zip("ABCDEFGHIJKLMN", [26, 10, 46, 12, 9, 17, 17, 13, 15, 40, 12, 46, 22, 12]):
        ws.column_dimensions[col].width = width
    wb.save(path)


def append_row(path: str, row: dict) -> None:
    if not os.path.exists(path):
        _new_book(path)
    wb = load_workbook(path)
    ws = wb["dataset"]
    values = [row.get(c, "") for c in C.EXCEL_COLUMNS] + [
        row.get("_trang_thai", ""),
        row.get("_url_nguon", ""),
        str(row.get("_phash", "")),
        row.get("_ma_du_an", "") or "",
    ]
    ws.append(values)
    fill = FILL.get(row.get("_trang_thai"))
    if fill:
        for cell in ws[ws.max_row]:
            cell.fill = fill
    wb.save(path)


def append_many(path: str, rows: list[dict]) -> None:
    """Ghi theo lo - nhanh hon nhieu so voi ghi tung dong."""
    if not rows:
        return
    if not os.path.exists(path):
        _new_book(path)
    wb = load_workbook(path)
    ws = wb["dataset"]
    for row in rows:
        ws.append(
            [row.get(c, "") for c in C.EXCEL_COLUMNS]
            + [
                row.get("_trang_thai", ""),
                row.get("_url_nguon", ""),
                str(row.get("_phash", "")),
                row.get("_ma_du_an", "") or "",
            ]
        )
        fill = FILL.get(row.get("_trang_thai"))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    wb.save(path)


def load_dedup(path: str) -> tuple[list[tuple[int, str]], set[str]]:
    """Doc lai pHash + ma du an da ghi -> chong trung GIUA CAC PHIEN chay.

    Khong co buoc nay thi chay `crawl` lan hai se tai lai va xu ly lai
    dung nhung du an da lam xong.
    """
    if not os.path.exists(path):
        return [], set()
    ws = load_workbook(path, read_only=True)["dataset"]
    base = len(C.EXCEL_COLUMNS)          # _trang_thai _url_nguon _phash _ma_du_an
    hashes: list[tuple[int, str]] = []
    codes: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < base + 4:
            continue
        state, phash_raw, code = row[base], row[base + 2], row[base + 3]
        if code:
            codes.add(str(code).strip())
        if state == "loai" or not phash_raw:
            continue                      # mau loai khong dung lam moc chong trung
        try:
            hashes.append((int(str(phash_raw).strip()), str(row[0] or "")))
        except ValueError:
            continue
    return hashes, codes


def summary(path: str) -> dict:
    if not os.path.exists(path):
        return {"dat": 0, "tam": 0, "loai": 0, "tong": 0, "ty_le_dat": 0.0}
    ws = load_workbook(path, read_only=True)["dataset"]
    counts = {"dat": 0, "tam": 0, "loai": 0}
    idx = len(C.EXCEL_COLUMNS)
    for row in ws.iter_rows(min_row=2, values_only=True):
        state = row[idx] if len(row) > idx else None
        if state in counts:
            counts[state] += 1
    counted = counts["dat"] + counts["tam"]
    counts["tong"] = sum(counts.values())
    counts["ty_le_dat"] = counts["dat"] / counted if counted else 0.0
    return counts


def print_progress(path: str) -> None:
    s = summary(path)
    pct = s["dat"] / C.QUOTA * 100
    bar = "#" * int(pct / 2.5) + "." * (40 - int(pct / 2.5))
    print(f"\n  [{bar}] {s['dat']}/{C.QUOTA}  ({pct:.1f}%)")
    print(f"  dat {s['dat']}  |  tam {s['tam']}  |  loai {s['loai']}")
    print(f"  ty le dat: {s['ty_le_dat']:.1%}  (yeu cau >= {C.TARGET_PASS_RATE:.0%})")
    for name, target in C.MILESTONES.items():
        mark = "x" if s["dat"] >= target else " "
        print(f"    [{mark}] {name:<6} {target}")
