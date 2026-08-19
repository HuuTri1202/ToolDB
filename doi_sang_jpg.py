#!/usr/bin/env python3
"""Doi moi anh trong thu muc ket qua sang .jpg va cap nhat cot ten_file.

Chay duoc nhieu lan, khong hong gi neu chay lai.
  python doi_sang_jpg.py            # xem truoc, khong sua gi
  python doi_sang_jpg.py --that     # thuc hien
"""
from __future__ import annotations

import argparse
import os
import sys

import cv2
from openpyxl import load_workbook

import main as M

CHAT_LUONG = 95


def doi_file(duong_dan: str, that: bool) -> str | None:
    """Doi mot anh sang .jpg. Tra ve ten file moi, hoac None neu bo qua."""
    goc, ext = os.path.splitext(duong_dan)
    ext = ext.lower()
    if ext == ".jpg":
        return None

    moi = goc + ".jpg"
    if os.path.exists(moi):
        print(f"  BO QUA (da co {os.path.basename(moi)})")
        return None
    if not that:
        return os.path.basename(moi)

    if ext == ".jpeg":                       # cung dinh dang, chi doi ten
        os.replace(duong_dan, moi)
        return os.path.basename(moi)

    img = cv2.imread(duong_dan)
    if img is None:
        print(f"  LOI doc {duong_dan}")
        return None
    cv2.imwrite(moi, img, [cv2.IMWRITE_JPEG_QUALITY, CHAT_LUONG])
    os.remove(duong_dan)
    return os.path.basename(moi)


def main() -> None:
    ap = argparse.ArgumentParser(description="Doi anh ket qua sang .jpg")
    ap.add_argument("--that", action="store_true", help="thuc hien that")
    ap.add_argument("--gom-png", action="store_true",
                    help="doi ca .png (mac dinh GIU vi png khong mat net)")
    args = ap.parse_args()

    duoi = [".webp", ".jpeg"] + ([".png"] if args.gom_png else [])
    doi_ten: dict[str, str] = {}

    for thu_muc in (M.IMG_DIR, M.TAM_DIR, M.LOAI_DIR):
        if not os.path.isdir(thu_muc):
            continue
        for ten in sorted(os.listdir(thu_muc)):
            if os.path.splitext(ten)[1].lower() not in duoi:
                continue
            print(f"{os.path.basename(thu_muc)}/{ten}")
            moi = doi_file(os.path.join(thu_muc, ten), args.that)
            if moi:
                doi_ten[ten] = moi
                print(f"  -> {moi}")

    if not doi_ten:
        print("Khong co file nao can doi.")
        return

    # Cap nhat cot ten_file trong Excel
    if os.path.exists(M.EXCEL) and args.that:
        wb = load_workbook(M.EXCEL)
        ws = wb["dataset"]
        sua = 0
        for row in ws.iter_rows(min_row=2):
            o = row[0]
            if o.value in doi_ten:
                o.value = doi_ten[o.value]
                sua += 1
        wb.save(M.EXCEL)
        print(f"\nExcel: cap nhat {sua} dong")

    print(f"\n{len(doi_ten)} file da doi." if args.that
          else f"\n{len(doi_ten)} file SE doi. Chay lai voi --that de thuc hien.")
    if not args.gom_png:
        print("Luu y: .png duoc GIU nguyen - JPEG lam nhoe net tuong mong.")
        print("Muon doi ca png thi them --gom-png.")


if __name__ == "__main__":
    sys.exit(main())
