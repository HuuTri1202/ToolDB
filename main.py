#!/usr/bin/env python3
"""Tool thu thap mat bang biet thu - nhom N3 Cong nghe Thong tin.

KHONG GOI AI, KHONG TON PHI. Toan bo pipeline chay bang regex + OpenCV.

  1. loc alt/caption bang regex        loai 60-70%
  2. thu URL ban goc                   cuu mau phan giai thap
  3. tai anh
  4. crop khung ban ve
  5. do phan giai (SAU crop)
  6. lineart + contrast + ty le canh
  7. dem duong thang ngang/doc         loai phoi canh va render
  8. watermark chim
  9. pHash chong trung
 10. ghi Excel de nguoi cham 5 cot con lai

Tool KHONG TU DAT mau nao thanh 'dat'. 16/20 tieu chi can mat nguoi nhin
(xem QC_HUONG_DAN.md). Mau song sot deu vao trang thai 'tam' cho QC.

Lenh:
  python main.py map                                       khong mang
  python main.py probe  --source neohouse --limit 10       chi doc HTML
  python main.py check  --dir ./anh_tai_tay                khong mang
  python main.py calib  --dir ./bo_test --labels nhan.csv  khong mang
  python main.py crawl  --source neohouse --limit 50       co mang
  python main.py report                                    khong mang
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import urllib.parse as up

import cv2

import assess
import config as C
import crawler
import criteria
import excel_out
import imgcheck

OUT_DIR = os.environ.get("N3_OUT", "./output")
IMG_DIR = os.path.join(OUT_DIR, "bietthu")     # chi mau 'dat' - thu muc de nop
TAM_DIR = os.path.join(OUT_DIR, "_tam")        # mau 'tam' - cho sua tay
LOAI_DIR = os.path.join(OUT_DIR, "_loai")      # mau 'loai' - de doi chieu khi hieu chinh
EXCEL = os.path.join(OUT_DIR, "bietthu_dataset.xlsx")

DEST_DIR = {"dat": IMG_DIR, "tam": TAM_DIR, "loai": LOAI_DIR}
TMP_DIR = os.path.join(OUT_DIR, "_tmp")


def _ensure_dirs() -> None:
    for d in (OUT_DIR, IMG_DIR, TAM_DIR, LOAI_DIR, TMP_DIR):
        os.makedirs(d, exist_ok=True)


def _next_index() -> int:
    """Quet CA hai thu muc: 'dat' va 'tam' deu tieu mot so thu tu."""
    prefix = C.HOUSE_TYPE + "_"
    nums: list[int] = []
    for d in (IMG_DIR, TAM_DIR):
        if not os.path.isdir(d):
            continue
        nums += [
            int(n[len(prefix) : len(prefix) + 3])
            for n in os.listdir(d)
            if n.startswith(prefix) and n[len(prefix) :][:3].isdigit()
        ]
    return max(nums, default=0) + 1


def _filename(idx: int, tang: int | None, ext: str) -> str:
    suffix = f"_t{tang}" if tang else ""
    return f"{C.HOUSE_TYPE}_{idx:03d}{suffix}{ext}"


def _luu_jpg(nguon: str, dich: str) -> None:
    """Chuyen anh sang .jpg khi luu.

    Nguon web tra ve ca .webp va .png; ho so nop yeu cau .jpg dong nhat.
    Anh da la jpg thi doi cho, khong ma hoa lai (tranh mat chat luong lan hai).
    """
    if os.path.splitext(nguon)[1].lower() in (".jpg", ".jpeg"):
        os.replace(nguon, dich)
        return

    img = cv2.imread(nguon)
    if img is None:                       # khong doc duoc -> giu nguyen file
        os.replace(nguon, dich)
        return
    cv2.imwrite(dich, img, [cv2.IMWRITE_JPEG_QUALITY, 95])
    os.remove(nguon)


def _loai_filename(ph: int, ext: str) -> str:
    """Mau 'loai' KHONG tieu so thu tu - dat ten theo pHash de khong ghi de nhau."""
    return f"loai_{ph:016x}{ext}"


def _load_so_can() -> dict[tuple, int]:
    """Doc lai so thu tu da cap cho tung CAN NHA tu Excel.

    Nho vay cao lai lan hai se gan tiep ban ve cua mot can vao dung so cu,
    thay vi tao ra mot so moi cho cung can do.
    """
    if not os.path.exists(EXCEL):
        return {}
    from openpyxl import load_workbook
    ws = load_workbook(EXCEL, read_only=True)["dataset"]
    base = len(C.EXCEL_COLUMNS)
    prefix = C.HOUSE_TYPE + "_"
    ra: dict[tuple, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ten = str(row[0] or "")
        if not ten.startswith(prefix) or len(row) < base + 2:
            continue
        so = os.path.splitext(ten)[0][len(prefix):][:3]
        if not so.isdigit():
            continue
        ten_goc = str(row[base + 1] or "").rsplit("/", 1)[-1]
        ra.setdefault((str(row[2] or ""), assess.khoa_can_nha(ten_goc)), int(so))
    return ra


def _ten_moi(so: int, nhan: str, ext: str, thu_muc: str) -> str:
    """Ten file cho mot ban ve; them chu cai neu can do da co ban ve cung tang."""
    goc = f"{C.HOUSE_TYPE}_{so:03d}{nhan}"
    if not os.path.exists(os.path.join(thu_muc, goc + ext)):
        return goc + ext
    for c in "bcdefgh":
        if not os.path.exists(os.path.join(thu_muc, f"{goc}{c}{ext}")):
            return f"{goc}{c}{ext}"
    return f"{goc}z{ext}"


def _load_dedup() -> imgcheck.DedupIndex:
    hashes, codes = excel_out.load_dedup(EXCEL)
    if hashes or codes:
        print(f"  nap lai chong trung tu Excel: {len(hashes)} pHash, "
              f"{len(codes)} ma du an")
    return imgcheck.DedupIndex(hashes, codes)


def _row(fname: str, ctx: dict, rep, v: dict, state: str) -> dict:
    tc = criteria.criteria_note(v, rep.flags)
    notes = [n for n in (v.get("ghi_chu"), *rep.flags, tc) if n]
    if ctx.get("so_tang"):
        notes.insert(0, f"nha {ctx['so_tang']} tang")
    return {
        "ten_file": fname,
        "loai_nha": C.HOUSE_TYPE,
        "nguon": ctx.get("project_url", ""),
        "co_noi_that": v.get("co_noi_that", "khong"),
        "do_net": v.get("do_net", "loai"),
        "dung_cho_REPLAN": v.get("dung_cho_REPLAN", "loai"),
        "dung_cho_GEPLAN": v.get("dung_cho_GEPLAN", "loai"),
        "so_phong_uoc": v.get("so_phong_uoc", 0),
        "co_boundary_ro": v.get("co_boundary_ro", "khong"),
        "ghi_chu": " · ".join(notes)[:250],
        "_trang_thai": state,
        "_url_nguon": ctx.get("url", ""),
        "_phash": rep.phash,
        "_ma_du_an": ctx.get("ma_du_an"),
    }


# --------------------------------------------------------------------- crawl
def cmd_crawl(args) -> None:
    _ensure_dirs()
    if args.source not in C.SOURCES:
        sys.exit(f"Nguon khong hop le. Co: {', '.join(C.SOURCES)}")

    dedup = _load_dedup()
    so_can = _load_so_can()
    idx = _next_index()
    rows: list[dict] = []
    stats = {"duyet": 0, "loc_text": 0, "loc_anh": 0, "trung": 0,
             "qua_loc": 0, "dat": 0, "tam": 0, "loai": 0}

    print(f"Tim trang du an tren {C.SOURCES[args.source]['name']} ...")
    projects = crawler.project_links(args.source, limit=args.limit)
    print(f"  -> {len(projects)} du an\n")

    for p_i, purl in enumerate(projects, 1):
        cands = crawler.candidate_images(purl)
        stats["duyet"] += 1
        print(f"[{p_i}/{len(projects)}] {purl.rsplit('/', 2)[-2][:52]}  "
              f"({len(cands)} ung vien)")

        if dedup.seen_code(cands[0].get("ma_du_an") if cands else None):
            print("    bo qua - ma du an da co")
            continue

        for ctx in cands:
            url = crawler.try_original(ctx["url"])
            ctx["url"] = url
            tai_ext = os.path.splitext(url.split("?")[0])[1].lower() or ".jpg"
            ext = ".jpg"          # ho so nop yeu cau .jpg dong nhat
            tmp = os.path.join(TMP_DIR, f"_tmp{tai_ext}")
            if not crawler.download(url, tmp):
                continue

            rep = imgcheck.inspect(tmp)
            if not rep.ok:
                stats["loc_anh"] += 1
                print(f"    - {rep.reason} ({rep.width}x{rep.height})")
                continue

            dup = dedup.seen_hash(rep.phash)
            if dup:
                stats["trung"] += 1
                print(f"    - trung voi {dup}")
                continue

            stats["qua_loc"] += 1
            v = assess.judge(ctx, rep)
            state = assess.gate(v, rep)
            stats[state] += 1

            if state == "loai":
                fname = _loai_filename(rep.phash, ext)
            else:
                # MOI CAN NHA mot so; hau to phan biet tang trong can do:
                #   bietthu_028.jpg  _t1  _t2  = tret, tang 1, tang 2 CUNG can
                # Khong the gom theo trang du an: nguon kieu trang tong hop co
                # mot trang chua hang chuc can, nen phai rut ten can tu ten file.
                ten_goc = url.split("?")[0].rsplit("/", 1)[-1]
                khoa = (purl, assess.khoa_can_nha(ten_goc))
                if khoa not in so_can:
                    so_can[khoa] = idx
                    idx += 1
                fname = _ten_moi(so_can[khoa], assess.nhan_tang(ten_goc),
                                 ext, DEST_DIR[state])
            _luu_jpg(tmp, os.path.join(DEST_DIR[state], fname))

            if state != "loai":
                dedup.add(rep.phash, fname)
                # idx da tang o tren khi gap CAN moi - khong tang theo tung anh
            rows.append(_row(fname, ctx, rep, v, state))
            print(f"    {state.upper():<5} {fname}  "
                  f"{rep.width}x{rep.height} duong_thang={rep.axis_lines}")

        if len(rows) >= 25:
            excel_out.append_many(EXCEL, rows)
            rows = []

    excel_out.append_many(EXCEL, rows)
    print("\n--- thong ke phien ---")
    for k, val in stats.items():
        print(f"  {k:<9} {val}")
    excel_out.print_progress(EXCEL)


# --------------------------------------------------------------------- probe
def _probe_url(url: str) -> None:
    """Do thu MOT trang bat ky chua co trong config.SOURCES.

    Dung de danh gia nguon moi truoc khi quyet dinh them vao config.
    """
    import collections
    import urllib.parse as up

    from bs4 import BeautifulSoup

    host = up.urlparse(url).netloc
    print(f"Do thu {host} - chi doc HTML, khong tai anh.\n")

    tin_hieu = crawler.ai_train_signal(url)
    if tin_hieu == "no":
        print("\n  Khong do tiep. Neu van muon dung nguon nay, phai xin phep")
        print("  chu site bang van ban truoc - hoi Mr. Phuc ve quy trinh.")
        return
    print(f"  Content-Signal ai-train: {tin_hieu or 'khong khai bao'}")

    if not crawler.allowed(url):
        print("  robots.txt CHAN URL nay voi User-Agent cua tool. Dung lai.")
        return

    resp = crawler.get(url)
    if resp is None:
        print("  Khong tai duoc trang (robots chan, timeout, hoac loi mang).")
        return

    soup = BeautifulSoup(resp.text, "html.parser")
    tieu_de = soup.title.get_text(strip=True) if soup.title else ""
    print(f"  Tieu de: {tieu_de[:66]}")

    # 1. Trang nay co phai trang du an khong (co anh mat bang truc tiep)?
    cands, st = crawler.scan_project(url)
    print(f"\n  Coi day la trang du an:")
    print(f"    the <img>            {st['the_img']}")
    print(f"    bi REJECT_TEXT chan  {st['bi_reject_text']}")
    print(f"    khong khop ACCEPT    {st['khong_khop_accept']}")
    print(f"    UNG VIEN             {st['ung_vien']}")
    for c in cands[:4]:
        print(f"      -> {c['url'].rsplit('/', 1)[-1][:62]}")

    # 2. Chuoi 'mat bang' co xuat hien trong HTML tho khong - phep thu 10 giay
    tho = resp.text.lower()
    dem = sum(tho.count(t) for t in ("mặt bằng", "mat bang", "mat-bang"))
    print(f"\n  Chuoi 'mat bang' trong HTML tho: {dem} lan")
    if dem == 0 and st["ung_vien"] == 0:
        print("    -> Trang khong nhac gi den mat bang. Kha nang cao la bo.")

    # 3. Goi y project_url_pattern tu cac link noi bo
    nhom = collections.Counter()
    for a in soup.select("a[href]"):
        p = up.urlparse(up.urljoin(url, a["href"]))
        if p.netloc != host:
            continue
        doan = [x for x in p.path.split("/") if x]
        if len(doan) >= 2:
            nhom[f"/{doan[0]}/"] += 1
    if nhom:
        print("\n  Tien to duong dan hay gap (goi y project_url_pattern):")
        for tien_to, n in nhom.most_common(5):
            print(f"    {tien_to:<34} {n} link")

    print("\n  Neu nguon dung duoc: them vao config.SOURCES roi chay")
    print("  `python main.py probe --source <ten>` de do ca listing.")


def cmd_probe(args) -> None:
    """Do thu nguon: chi doc HTML. Khong tai anh, khong goi AI, khong ton phi.

    Chay truoc `crawl` de biet selector con dung khong va loc alt text
    con qua chat hay khong.
    """
    if args.url:
        return _probe_url(args.url)

    if args.source not in C.SOURCES:
        sys.exit(f"Nguon khong hop le. Co: {', '.join(C.SOURCES)}")

    src = C.SOURCES[args.source]
    print(f"Do thu {src['name']} - chi doc HTML, khong tai anh, khong goi AI.\n")
    projects = crawler.project_links(args.source, limit=args.limit)
    print(f"  bat duoc {len(projects)} trang du an\n")
    if not projects:
        print("  0 trang du an. Kiem tra theo thu tu:")
        print(f"    1. mo thu 1 URL listing tren trinh duyet: {src['listing'][0]}")
        print(f"    2. sua SOURCES['{args.source}']['project_url_pattern'] "
              f"(dang la {src['project_url_pattern']!r})")
        print("    3. robots.txt co dang chan User-Agent cua ta khong")
        return

    tong = {"the_img": 0, "sai_duoi": 0, "bi_reject_text": 0,
            "khong_khop_accept": 0, "ung_vien": 0, "loi_tai_trang": 0}
    co_meta = 0

    for i, purl in enumerate(projects, 1):
        cands, st = crawler.scan_project(purl)
        for k in tong:
            tong[k] += st.get(k, 0)
        ma = cands[0].get("ma_du_an") if cands else None
        tang = cands[0].get("so_tang") if cands else None
        if ma or tang:
            co_meta += 1
        print(f"[{i}/{len(projects)}] {purl.rsplit('/', 2)[-2][:46]:<46} "
              f"img={st['the_img']:>3}  ung_vien={st['ung_vien']:>2}  "
              f"tang={tang or '-'} ma={ma or '-'}")

    xet = tong["the_img"] - tong["sai_duoi"]
    ty_le = tong["ung_vien"] / xet if xet else 0.0
    print("\n--- tong ket ---")
    print(f"  the <img>            {tong['the_img']}")
    print(f"  sai duoi file        {tong['sai_duoi']}")
    print(f"  bi REJECT_TEXT chan  {tong['bi_reject_text']}")
    print(f"  khong khop ACCEPT    {tong['khong_khop_accept']}")
    print(f"  UNG VIEN             {tong['ung_vien']}   ({ty_le:.0%} so anh duoc xet)")
    print(f"  trang co metadata    {co_meta}/{len(projects)}")
    if tong["loi_tai_trang"]:
        print(f"  loi tai trang        {tong['loi_tai_trang']}")

    print()
    if tong["ung_vien"] == 0:
        print("  0 ung vien. Mo mot trang du an, xem alt text cua anh mat bang")
        print("  roi them tu khoa vao config.ACCEPT_TEXT.")
    elif ty_le < 0.05:
        print("  Ty le ung vien rat thap - co the ACCEPT_TEXT dang qua chat,")
        print("  hoac trang dung lazy-load (anh nam o data-src, khong phai src).")
    else:
        uoc = tong["ung_vien"] / max(len(projects), 1)
        print(f"  Uoc tinh ~{uoc:.1f} ung vien/du an -> khoang "
              f"{uoc * 100:.0f} anh phai tai va loc neu crawl 100 du an.")
    print("  Buoc tiep theo: python main.py crawl --source "
          f"{args.source} --limit 5")


# --------------------------------------------------------------------- danhso
def cmd_danhso(args) -> None:
    """Danh so lai lien tuc 001..N, giu nguyen hau to tang.

    Sau nhieu phien cao va vai lan don dep, day so bi thung (thieu 26, 119...).
    Lenh nay dong lai cho lien tuc de de kiem va de nop.

    Doi ten qua HAI BUOC de khong bao gio de len file dang ton tai.
    """
    if not os.path.exists(EXCEL):
        sys.exit(f"Khong thay {EXCEL}")

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL)
    ws = wb["dataset"]
    prefix = C.HOUSE_TYPE + "_"

    # Gom file theo thu muc, sap theo so hien tai de giu nguyen thu tu cu
    viec: list[tuple[str, str, str]] = []      # (thu_muc, ten_cu, hau_to)
    for d in (IMG_DIR, TAM_DIR):
        if not os.path.isdir(d):
            continue
        ten_ds = [t for t in os.listdir(d) if t.startswith(prefix)]

        def khoa(t: str) -> int:
            so = t[len(prefix):][:3]
            return int(so) if so.isdigit() else 9999

        for t in sorted(ten_ds, key=khoa):
            goc, ext = os.path.splitext(t)
            duoi = goc[len(prefix) + 3:]        # "_t2" hoac ""
            viec.append((d, t, duoi + ext))

    doi: dict[str, str] = {}
    for i, (d, cu, duoi) in enumerate(viec, 1):
        moi = f"{prefix}{i:03d}{duoi}"
        if moi != cu:
            doi[cu] = moi

    print(f"{len(viec)} file | {len(doi)} file se doi so")
    for cu, moi in list(doi.items())[:10]:
        print(f"  {cu:<26} -> {moi}")
    if len(doi) > 10:
        print(f"  ... con {len(doi) - 10} file")

    if not args.that:
        print("\nChay lai voi --that de thuc hien.")
        return

    # Buoc 1: doi sang ten tam, tranh de len file dang co
    tam: list[tuple[str, str, str]] = []
    for i, (d, cu, duoi) in enumerate(viec, 1):
        moi = f"{prefix}{i:03d}{duoi}"
        if moi == cu:
            continue
        t = os.path.join(d, f"__dsl_{i:04d}{os.path.splitext(cu)[1]}")
        os.replace(os.path.join(d, cu), t)
        tam.append((t, os.path.join(d, moi), cu))

    # Buoc 2: tu ten tam sang ten cuoi
    for t, dich, _ in tam:
        os.replace(t, dich)

    sua = 0
    for row in ws.iter_rows(min_row=2):
        v = str(row[0].value or "")
        if v in doi:
            row[0].value = doi[v]
            sua += 1
    wb.save(EXCEL)
    print(f"\nDa doi {len(tam)} file | Excel cap nhat {sua} dong")


# --------------------------------------------------------------------- retag
def cmd_retag(args) -> None:
    """Gan lai hau to _t1/_t2 cho anh DA CO, khong phai cao lai.

    Doc lai so tang tu cot _url_nguon (URL anh goc) trong Excel, roi doi ten
    file tren dia va cap nhat cot ten_file cho khop.
    """
    if not os.path.exists(EXCEL):
        sys.exit(f"Khong thay {EXCEL}")

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL)
    ws = wb["dataset"]
    base = len(C.EXCEL_COLUMNS)              # _trang_thai _url_nguon ...

    doi, giu, thieu = 0, 0, 0
    for row in ws.iter_rows(min_row=2):
        o_ten, o_url = row[0], row[base + 1]
        ten = str(o_ten.value or "")
        if not ten.startswith(C.HOUSE_TYPE + "_"):
            continue                          # mau 'loai' dat ten theo pHash

        # CHI doc URL anh goc. Khong doc ghi_chu: do la du lieu tool tu suy ra,
        # doc lai se nhan ban loi cu (mot mau tung bi ghi nham "tang 7").
        tang = assess.so_tang({"url": str(o_url.value or "")})
        goc, ext = os.path.splitext(ten)
        so = goc[len(C.HOUSE_TYPE) + 1:][:3]

        if not tang:
            # Khong suy ra duoc tu URL -> GIU nguyen ten cu. Nhieu mau co hau to
            # nho alt/caption luc cao, ma Excel khong luu alt, nen xoa la mat.
            giu += 1
            continue
        moi = f"{C.HOUSE_TYPE}_{so}_t{tang}{ext}"
        if moi == ten:
            giu += 1
            continue

        cu_dd = moi_dd = None
        for d in (IMG_DIR, TAM_DIR, LOAI_DIR):
            if os.path.exists(os.path.join(d, ten)):
                cu_dd, moi_dd = os.path.join(d, ten), os.path.join(d, moi)
                break
        if cu_dd is None:
            thieu += 1
            print(f"  thieu file tren dia: {ten}")
            continue
        if os.path.exists(moi_dd):
            print(f"  BO QUA (trung ten): {ten} -> {moi}")
            continue

        print(f"  {ten:<26} -> {moi}")
        if args.that:
            os.replace(cu_dd, moi_dd)
            o_ten.value = moi
        doi += 1

    if args.that:
        wb.save(EXCEL)
    print(f"\n{doi} file {'da' if args.that else 'SE'} doi ten | "
          f"{giu} giu nguyen | {thieu} thieu file")
    if not args.that:
        print("Chay lai voi --that de thuc hien.")


# --------------------------------------------------------------------- screen
def cmd_screen(args) -> None:
    """Sang NHIEU nguon mot luot: nguon nao dang cao, nguon nao bo.

    Voi moi URL bao 4 tin hieu, khong tai anh nao:
      ai-train  - chu site co tu choi cho huan luyen mo hinh khong
      mat_bang  - so lan chuoi 'mat bang' xuat hien trong HTML tho
      ung_vien  - so anh qua duoc bo loc alt/caption/ten file
      ket luan  - CAO / BO / CHAN
    """
    urls = list(args.urls or [])
    if args.file:
        with open(args.file, encoding="utf-8-sig") as fh:
            urls += [d.strip() for d in fh if d.strip() and not d.startswith("#")]
    if not urls:
        sys.exit("Can --urls <url...> hoac --file <danh_sach.txt>")

    print(f"Sang {len(urls)} nguon - chi doc HTML, khong tai anh.\n")
    print(f"{'nguon':<26}{'ai-train':<11}{'mat_bang':>9}{'ung_vien':>10}  ket luan")
    print("-" * 78)

    nen_cao = []
    for u in urls:
        host = up.urlparse(u).netloc or u
        u = u if u.startswith("http") else f"https://{u}"

        tin = crawler.ai_train_signal(u)
        if tin == "no":
            print(f"{host:<26}{'no':<11}{'-':>9}{'-':>10}  CHAN - chu site tu choi")
            continue

        resp = crawler.get(u)
        if resp is None:
            print(f"{host:<26}{tin or '-':<11}{'-':>9}{'-':>10}  BO - khong tai duoc")
            continue

        tho = resp.text.lower()
        mb = sum(tho.count(t) for t in ("mặt bằng", "mat bang", "mat-bang"))
        cands, st = crawler.scan_project(u)
        n = st["ung_vien"]

        if n >= 3 or (mb >= 5 and n >= 1):
            ket, nen_cao = "CAO - dang thu", nen_cao + [host]
        elif mb == 0:
            ket = "BO - khong nhac mat bang"
        elif n == 0:
            ket = "NGO - co chu nhung 0 anh khop"
        else:
            ket = "NGO - it ung vien"
        print(f"{host:<26}{tin or 'khong':<11}{mb:>9}{n:>10}  {ket}")

    print("\n" + "-" * 78)
    if nen_cao:
        print(f"Nen dao sau {len(nen_cao)} nguon: {', '.join(nen_cao)}")
        print("Buoc tiep: python main.py probe --url \"<url cu the>\" de xem ky.")
    else:
        print("Khong nguon nao dat nguong. Thu URL khac cua cung site truoc khi bo.")
    print("Luu y: 'NGO' khong co nghia la bo - co the URL nay la trang tong hop,")
    print("hay thu mot trang du an cu the cua site do.")


# --------------------------------------------------------------------- check
def cmd_check(args) -> None:
    _ensure_dirs()
    files = sorted(
        f for f in os.listdir(args.dir) if f.lower().endswith(C.ALLOWED_EXT)
    )
    if not files:
        sys.exit("Khong tim thay anh trong thu muc.")

    print(f"  {len(files)} anh - loc bang OpenCV, khong goi AI, khong ton phi.")
    dedup = _load_dedup()
    rows: list[dict] = []
    idx = _next_index()

    for f in files:
        path = os.path.join(args.dir, f)
        rep = imgcheck.inspect(path)
        if not rep.ok:
            print(f"  LOAI  {f:<34} {rep.reason} ({rep.width}x{rep.height})")
            continue
        if dedup.seen_hash(rep.phash):
            print(f"  LOAI  {f:<34} trung lap")
            continue

        v = assess.judge({"title": f, "alt": f}, rep)
        state = assess.gate(v, rep)
        ext = os.path.splitext(f)[1].lower()
        if state == "loai":
            fname = _loai_filename(rep.phash, ext)
        else:
            fname = _filename(idx, v.get("so_tang_tren_anh") or None, ext)
            dedup.add(rep.phash, fname)
            idx += 1
        rows.append(_row(fname, {"project_url": "thu_cong", "url": path}, rep, v, state))
        print(f"  {state.upper():<5} {f:<34} {rep.width}x{rep.height} "
              f"duong_thang={rep.axis_lines} do_net={v.get('do_net')}")

    excel_out.append_many(EXCEL, rows)
    excel_out.print_progress(EXCEL)


# --------------------------------------------------------------------- calib
def cmd_calib(args) -> None:
    """So ket qua tool voi nhan tay -> ma tran hieu chinh.

    File nhan CSV: ten_file,nhan   (nhan = dat|tam|loai)
    """
    labels: dict[str, str] = {}
    with open(args.labels, encoding="utf-8-sig") as fh:
        for r in csv.reader(fh):
            if len(r) < 2 or not r[0].strip():
                continue
            nhan = r[1].strip().lower()
            if nhan not in ("dat", "tam", "loai"):
                continue          # bo qua dong tieu de va dong ghi chu
            labels[r[0].strip()] = nhan
    if not labels:
        sys.exit("File nhan khong co dong hop le. Mau: ten_file,nhan (dat|tam|loai)")

    matrix = {(a, b): 0 for a in ("dat", "tam", "loai") for b in ("dat", "tam", "loai")}
    for f, human in labels.items():
        path = os.path.join(args.dir, f)
        if not os.path.exists(path):
            print(f"  thieu file: {f}")
            continue
        rep = imgcheck.inspect(path)
        if not rep.ok:
            tool = "tam" if "thu_lai_ban_goc" in rep.flags else "loai"
        else:
            tool = assess.gate(assess.judge({"title": f, "alt": f}, rep), rep)
        matrix[(human, tool)] += 1
        flag = "" if human == tool else "   <-- LECH"
        print(f"  {f:<34} nguoi={human:<5} tool={tool:<5}{flag}")

    print("\n            tool_dat  tool_tam  tool_loai")
    for h in ("dat", "tam", "loai"):
        print(f"  nguoi_{h:<5} "
              f"{matrix[(h,'dat')]:>8}  {matrix[(h,'tam')]:>8}  {matrix[(h,'loai')]:>9}")

    # Tool khong bao gio tra 'dat' (chi nguoi dat duoc), nen cot tool_dat luon 0.
    c = matrix[("dat", "loai")] + matrix[("tam", "loai")]
    g = matrix[("loai", "tam")]
    total = sum(matrix.values())
    print(f"\n  C = mau dung duoc bi loai oan (giet quota):   {c}")
    print(f"  G = mau rac lot sang khau cham tay:          {g}")
    print(f"  C+G = {c + g}/{total} - can C duoi 3/50 truoc khi mo van.")
    if c:
        print("  -> Uu tien sua C: ha MIN_AXIS_LINES, noi ASPECT_MIN/MAX,")
        print("     hoac bot tu khoa trong config.REJECT_TEXT.")
    if g:
        print("  -> G chi lam ton cong cham tay, khong giet quota. Chap nhan duoc.")


# --------------------------------------------------------------------- report
def cmd_report(_) -> None:
    excel_out.print_progress(EXCEL)


def main() -> None:
    ap = argparse.ArgumentParser(description="Tool dataset mat bang bietthu - N3")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("probe", help="do thu nguon: chi doc HTML, khong tai anh")
    p.add_argument("--source", default="neohouse")
    p.add_argument("--url", help="do thu MOT URL bat ky chua co trong config")
    p.add_argument("--limit", type=int, default=10, help="so trang du an de do thu")
    p.set_defaults(func=cmd_probe)

    p = sub.add_parser("crawl", help="thu thap tu nguon da cau hinh")
    p.add_argument("--source", default="neohouse")
    p.add_argument("--limit", type=int, default=50, help="so trang du an toi da")
    p.set_defaults(func=cmd_crawl)

    p = sub.add_parser("danhso", help="danh so lai lien tuc 001..N")
    p.add_argument("--that", action="store_true", help="thuc hien that")
    p.set_defaults(func=cmd_danhso)

    p = sub.add_parser("retag", help="gan lai hau to _t1/_t2 cho anh da co")
    p.add_argument("--that", action="store_true", help="thuc hien that")
    p.set_defaults(func=cmd_retag)

    p = sub.add_parser("screen", help="sang nhieu nguon mot luot truoc khi cao")
    p.add_argument("--urls", nargs="+", help="danh sach URL cach nhau bang dau cach")
    p.add_argument("--file", help="file txt, moi dong mot URL")
    p.set_defaults(func=cmd_screen)

    p = sub.add_parser("check", help="cham anh da co san trong thu muc")
    p.add_argument("--dir", required=True)
    p.set_defaults(func=cmd_check)

    p = sub.add_parser("calib", help="so tool voi nhan tay tren bo test")
    p.add_argument("--dir", required=True)
    p.add_argument("--labels", required=True)
    p.set_defaults(func=cmd_calib)

    p = sub.add_parser("report", help="bao cao tien do quota")
    p.set_defaults(func=cmd_report)

    p = sub.add_parser("map", help="bang doi chieu tieu chi II.4/II.5 -> code")
    p.set_defaults(func=lambda _: criteria.coverage_report())

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
