#!/usr/bin/env python3
"""Doi soat Excel voi anh tren dia, dung pHash lam khoa.

pHash khong doi du file bi doi ten hay doi dinh dang, nen no la khoa dang tin
duy nhat khi Excel va thu muc lech nhau.

  python doi_soat.py            # chi bao cao
  python doi_soat.py --that     # sua Excel cho khop dia
"""
from __future__ import annotations

import argparse
import os
import sys

import cv2
from openpyxl import load_workbook

import config as C
import imgcheck
import main as M


def quet_dia() -> dict[int, str]:
    """pHash -> ten file, cho moi anh dang co tren dia.

    Phai lap DUNG chuoi xu ly cua imgcheck.inspect(): doc anh -> crop khung
    ban ve -> xam -> phash. Bo buoc crop thi hash khac hoan toan.
    """
    ra: dict[int, str] = {}
    for d in (M.IMG_DIR, M.TAM_DIR, M.LOAI_DIR):
        if not os.path.isdir(d):
            continue
        for ten in sorted(os.listdir(d)):
            if not ten.lower().endswith(C.ALLOWED_EXT):
                continue
            img = cv2.imread(os.path.join(d, ten))
            if img is None:
                continue
            cropped, _ = imgcheck.crop_drawing_area(img)
            gray = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
            ra.setdefault(imgcheck.phash(gray), ten)
    return ra


def main() -> None:
    ap = argparse.ArgumentParser(description="Doi soat Excel voi thu muc anh")
    ap.add_argument("--that", action="store_true", help="sua Excel that")
    ap.add_argument("--xoa-mo-coi", action="store_true",
                    help="xoa dong Excel khong con file tren dia")
    args = ap.parse_args()

    if not os.path.exists(M.EXCEL):
        sys.exit(f"Khong thay {M.EXCEL}")

    tren_dia = quet_dia()
    print(f"Tren dia: {len(tren_dia)} anh duy nhat theo pHash")

    wb = load_workbook(M.EXCEL)
    ws = wb["dataset"]
    base = len(C.EXCEL_COLUMNS)

    khop, sua, mo_coi = 0, 0, []
    for row in ws.iter_rows(min_row=2):
        o_ten, o_ph = row[0], row[base + 2]
        ten = str(o_ten.value or "")
        try:
            ph = int(str(o_ph.value or "").strip())
        except ValueError:
            ph = None

        that = tren_dia.get(ph) if ph is not None else None
        if that is None:
            mo_coi.append((o_ten.row, ten))
        elif that == ten:
            khop += 1
        else:
            print(f"  lech: Excel={ten:<24} dia={that}")
            if args.that:
                o_ten.value = that
            sua += 1

    print(f"\nkhop: {khop} | lech ten: {sua} | mo coi (khong con file): {len(mo_coi)}")
    for r, t in mo_coi[:12]:
        print(f"    dong {r}: {t}")
    if len(mo_coi) > 12:
        print(f"    ... con {len(mo_coi) - 12} dong nua")

    if args.that and args.xoa_mo_coi and mo_coi:
        for r, _ in sorted(mo_coi, key=lambda x: -x[0]):
            ws.delete_rows(r)
        print(f"\nDa xoa {len(mo_coi)} dong mo coi")

    if args.that:
        wb.save(M.EXCEL)
        print("Da luu Excel")
    else:
        print("\nChay lai voi --that de sua (them --xoa-mo-coi de xoa dong thua)")


if __name__ == "__main__":
    sys.exit(main())
